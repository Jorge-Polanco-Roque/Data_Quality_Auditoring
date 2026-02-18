"""
Excel Report Generator — exporta reporte a .xlsx con pestañas formateadas.
Pestañas: Resumen, Por Columna, Issues Detallados, Filas Flaggeadas, Profiling.
Usa openpyxl con formato condicional.
"""

from typing import Dict, Optional

import pandas as pd

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


# Colores por severidad
SEVERITY_FILLS = {
    "CRITICAL": "FF0000",
    "HIGH": "FF8C00",
    "MEDIUM": "FFD700",
    "LOW": "90EE90",
    "INFO": "87CEEB",
    "PASS": "FFFFFF",
}

GRADE_FILLS = {
    "A": "00CC00",
    "B": "87CEEB",
    "C": "FFD700",
    "D": "FF8C00",
    "F": "FF0000",
}


def generate_excel(
    report: Dict,
    output_path: str,
    flagged_df: Optional[pd.DataFrame] = None,
):
    """Genera reporte Excel con múltiples pestañas."""
    if not HAS_OPENPYXL:
        raise ImportError("Se requiere openpyxl para generar Excel: pip install openpyxl")

    wb = Workbook()

    _build_summary_sheet(wb, report)
    _build_columns_sheet(wb, report)
    _build_issues_sheet(wb, report)
    _build_profiling_sheet(wb, report)

    if flagged_df is not None and len(flagged_df) > 0:
        _build_flagged_sheet(wb, flagged_df)

    # Remover la hoja por defecto si existe y hay otras
    if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
        del wb["Sheet"]

    wb.save(output_path)


def _style_header(ws, row, n_cols):
    """Aplica estilo a la fila de encabezado."""
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2F4F4F", end_color="2F4F4F", fill_type="solid")
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)


def _auto_width(ws):
    """Ajusta el ancho de columnas automáticamente."""
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 3, 50)


def _build_summary_sheet(wb, report):
    """Pestaña Resumen."""
    ws = wb.active
    ws.title = "Resumen"

    meta = report["report_metadata"]
    summary = report["dataset_summary"]

    # Título
    ws["A1"] = "DATA QUALITY AUDIT REPORT"
    ws["A1"].font = Font(bold=True, size=16)
    ws.merge_cells("A1:D1")

    # Info del archivo
    info_rows = [
        ("Archivo", meta["file_analyzed"]),
        ("Filas", f"{meta['total_rows']:,}"),
        ("Columnas", str(meta["total_columns"])),
        ("Encoding", meta["encoding"]),
        ("Generado", meta["generated_at"][:19]),
        ("", ""),
        ("Health Score", f"{summary['health_score']}/100"),
        ("Grado", summary["health_grade"]),
        ("Total Issues", str(summary["total_issues"])),
    ]

    for i, (label, value) in enumerate(info_rows, 3):
        ws.cell(row=i, column=1, value=label).font = Font(bold=True)
        ws.cell(row=i, column=2, value=value)

    # Color del grado
    grade = summary["health_grade"]
    grade_cell = ws.cell(row=10, column=2)
    if grade in GRADE_FILLS:
        grade_cell.fill = PatternFill(start_color=GRADE_FILLS[grade], end_color=GRADE_FILLS[grade], fill_type="solid")

    # Severidades
    row = 13
    ws.cell(row=row, column=1, value="Severidad").font = Font(bold=True)
    ws.cell(row=row, column=2, value="Cantidad").font = Font(bold=True)
    for level in ("CRITICAL", "HIGH", "MEDIUM", "LOW", "INFO"):
        row += 1
        count = summary["issues_by_severity"].get(level, 0)
        ws.cell(row=row, column=1, value=level)
        ws.cell(row=row, column=2, value=count)
        if count > 0 and level in SEVERITY_FILLS:
            ws.cell(row=row, column=1).fill = PatternFill(
                start_color=SEVERITY_FILLS[level], end_color=SEVERITY_FILLS[level], fill_type="solid"
            )

    # Trend
    trend = report.get("quality_trend")
    if trend and trend.get("previous_runs", 0) > 0:
        row += 2
        ws.cell(row=row, column=1, value="Tendencia de Calidad").font = Font(bold=True, size=12)
        row += 1
        ws.cell(row=row, column=1, value="Corridas anteriores").font = Font(bold=True)
        ws.cell(row=row, column=2, value=trend["previous_runs"])
        row += 1
        ws.cell(row=row, column=1, value="Tendencia").font = Font(bold=True)
        ws.cell(row=row, column=2, value=trend.get("trend_description", "N/A"))
        if "delta_vs_previous" in trend:
            row += 1
            ws.cell(row=row, column=1, value="Delta vs anterior").font = Font(bold=True)
            ws.cell(row=row, column=2, value=trend["delta_vs_previous"])

    _auto_width(ws)


def _build_columns_sheet(wb, report):
    """Pestaña Por Columna."""
    ws = wb.create_sheet("Por Columna")

    headers = ["Columna", "Tipo Semántico", "Score", "Grado", "% Nulos", "Únicos", "Checks", "Fallidos"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=1, column=i, value=h)
    _style_header(ws, 1, len(headers))

    row = 2
    profiles = report.get("column_profiles", {})
    for col_name, profile in sorted(profiles.items(), key=lambda x: x[1].get("health_score", 100)):
        ws.cell(row=row, column=1, value=col_name)
        ws.cell(row=row, column=2, value=profile.get("semantic_type", ""))
        ws.cell(row=row, column=3, value=profile.get("health_score", 0))
        grade = profile.get("health_grade", "?")
        ws.cell(row=row, column=4, value=grade)
        if grade in GRADE_FILLS:
            ws.cell(row=row, column=4).fill = PatternFill(
                start_color=GRADE_FILLS[grade], end_color=GRADE_FILLS[grade], fill_type="solid"
            )
        ws.cell(row=row, column=5, value=f"{profile.get('null_pct', 0):.1%}")
        ws.cell(row=row, column=6, value=profile.get("n_unique", 0))
        ws.cell(row=row, column=7, value=profile.get("checks_run", 0))
        ws.cell(row=row, column=8, value=profile.get("checks_failed", 0))
        row += 1

    _auto_width(ws)


def _build_issues_sheet(wb, report):
    """Pestaña Issues Detallados."""
    ws = wb.create_sheet("Issues Detallados")

    headers = ["Columna", "Check ID", "Severidad", "Mensaje", "Valor", "Umbral", "Afectados", "% Afectados"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=1, column=i, value=h)
    _style_header(ws, 1, len(headers))

    row = 2
    for col_name, profile in report.get("column_profiles", {}).items():
        for issue in profile.get("issues", []):
            ws.cell(row=row, column=1, value=col_name)
            ws.cell(row=row, column=2, value=issue.get("check_id", ""))
            sev = issue.get("severity", "")
            ws.cell(row=row, column=3, value=sev)
            if sev in SEVERITY_FILLS:
                ws.cell(row=row, column=3).fill = PatternFill(
                    start_color=SEVERITY_FILLS[sev], end_color=SEVERITY_FILLS[sev], fill_type="solid"
                )
            ws.cell(row=row, column=4, value=str(issue.get("message", ""))[:200])
            ws.cell(row=row, column=5, value=issue.get("value", ""))
            ws.cell(row=row, column=6, value=issue.get("threshold", ""))
            ws.cell(row=row, column=7, value=issue.get("affected_count", 0))
            ws.cell(row=row, column=8, value=f"{issue.get('affected_pct', 0):.2%}")
            row += 1

    _auto_width(ws)


def _build_profiling_sheet(wb, report):
    """Pestaña Profiling estadístico."""
    ws = wb.create_sheet("Profiling")

    stats = report.get("statistical_summary", {})

    # Numéricas
    num_stats = stats.get("numeric_columns", {})
    if num_stats:
        ws.cell(row=1, column=1, value="COLUMNAS NUMÉRICAS").font = Font(bold=True, size=12)
        headers = ["Columna", "Media", "Mediana", "Std", "Min", "Max", "Skewness", "Kurtosis", "Outliers IQR", "Outliers Z"]
        for i, h in enumerate(headers, 1):
            ws.cell(row=2, column=i, value=h)
        _style_header(ws, 2, len(headers))

        row = 3
        for col, ns in num_stats.items():
            ws.cell(row=row, column=1, value=col)
            ws.cell(row=row, column=2, value=ns.get("mean", ""))
            ws.cell(row=row, column=3, value=ns.get("median", ""))
            ws.cell(row=row, column=4, value=ns.get("std", ""))
            ws.cell(row=row, column=5, value=ns.get("min", ""))
            ws.cell(row=row, column=6, value=ns.get("max", ""))
            ws.cell(row=row, column=7, value=ns.get("skewness", ""))
            ws.cell(row=row, column=8, value=ns.get("kurtosis", ""))
            ws.cell(row=row, column=9, value=ns.get("outlier_count_iqr", 0))
            ws.cell(row=row, column=10, value=ns.get("outlier_count_zscore", 0))
            row += 1
        row += 1

    # Categóricas
    cat_stats = stats.get("categorical_columns", {})
    if cat_stats:
        ws.cell(row=row, column=1, value="COLUMNAS CATEGÓRICAS").font = Font(bold=True, size=12)
        row += 1
        cat_headers = ["Columna", "Únicos", "Valor Top", "Freq Top", "Categorías Raras"]
        for i, h in enumerate(cat_headers, 1):
            ws.cell(row=row, column=i, value=h)
        _style_header(ws, row, len(cat_headers))
        row += 1

        for col, cs in cat_stats.items():
            ws.cell(row=row, column=1, value=col)
            ws.cell(row=row, column=2, value=cs.get("n_unique", 0))
            ws.cell(row=row, column=3, value=cs.get("top_value", ""))
            ws.cell(row=row, column=4, value=f"{cs.get('top_freq', 0):.1%}")
            rare = cs.get("rare_categories", [])
            ws.cell(row=row, column=5, value=", ".join(rare[:5]) if rare else "ninguna")
            row += 1

    # Profiling extendido
    profiling = report.get("column_profiling", {})
    if profiling:
        row += 2
        ws.cell(row=row, column=1, value="PROFILING DETALLADO").font = Font(bold=True, size=12)
        row += 1
        prof_headers = ["Columna", "p5", "p25", "p50", "p75", "p95", "IQR", "CV", "Top 5 Valores"]
        for i, h in enumerate(prof_headers, 1):
            ws.cell(row=row, column=i, value=h)
        _style_header(ws, row, len(prof_headers))
        row += 1

        for col, p in profiling.items():
            ws.cell(row=row, column=1, value=col)
            pcts = p.get("percentiles", {})
            ws.cell(row=row, column=2, value=pcts.get("p5", ""))
            ws.cell(row=row, column=3, value=pcts.get("p25", ""))
            ws.cell(row=row, column=4, value=pcts.get("p50", ""))
            ws.cell(row=row, column=5, value=pcts.get("p75", ""))
            ws.cell(row=row, column=6, value=pcts.get("p95", ""))
            ws.cell(row=row, column=7, value=p.get("iqr", ""))
            ws.cell(row=row, column=8, value=p.get("cv", ""))
            top_vals = p.get("top_values", [])
            ws.cell(row=row, column=9, value=", ".join(f"{v[0]}({v[1]})" for v in top_vals[:5]) if top_vals else "")
            row += 1

    _auto_width(ws)


def _build_flagged_sheet(wb, flagged_df):
    """Pestaña Filas Flaggeadas."""
    ws = wb.create_sheet("Filas Flaggeadas")

    headers = list(flagged_df.columns)
    for i, h in enumerate(headers, 1):
        ws.cell(row=1, column=i, value=h)
    _style_header(ws, 1, len(headers))

    for row_idx, row_data in enumerate(flagged_df.itertuples(index=False), 2):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=str(value)[:200])

        # Color de severidad
        sev_col = headers.index("severity") + 1 if "severity" in headers else None
        if sev_col:
            sev = str(ws.cell(row=row_idx, column=sev_col).value)
            if sev in SEVERITY_FILLS:
                ws.cell(row=row_idx, column=sev_col).fill = PatternFill(
                    start_color=SEVERITY_FILLS[sev], end_color=SEVERITY_FILLS[sev], fill_type="solid"
                )

        if row_idx > 1001:  # Limitar a 1000 filas
            break

    _auto_width(ws)
